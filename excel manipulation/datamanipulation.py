import openpyxl

# wb = Workbook()
# ws = wb.active
# ws1 = wb.create_sheet("Mysheet")


# source = wb.active
# target = wb.copy_worksheet(source)

# print(wb.sheetnames)
# wb.save('balances.xlsx')
wb = openpyxl.load_workbook(filename="Intozi-Analytics.xlsx")
sheet_ranges = wb["FRS-Report"]
print(sheet_ranges["A8:K15"])
