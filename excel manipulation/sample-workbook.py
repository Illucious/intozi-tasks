from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()

ws = wb.active
# ws1 = wb.create_sheet('crowd-analytics')
# ws2 = wb.create_sheet('PPE-KITS')
# ws3 = wb.create_sheet('VIDS')
# ws4 = wb.create_sheet('VSDS')
# ws5 = wb.create_sheet('VSDS-Classification')
ws.title = "FRS-Report"

print(wb.sheetnames)
wb.save('sample.xlsx')

data = [
    ['Record no.','Record_id', 'Detection_Date', 'Detection_Time', 'location', 'device_name', 'Direction', 'person name', 'person type', 'registration-date', 'registration_id'],
    [1,3456,20-10-2023,'11:45:00', 'zone 4', 'entrance-1', 'entry', 'NA', 'Guest', 'NA','NA'],
    [2,1198,20-10,2023,'12:00:00', 'zome-1', 'exit-2', "Vishesh", 'Registered', '22-05-22', 'IT']

    
]

# add column headings. NB. these must be strings
# ws.append(["Fruit", "2011", "2012", "2013", "2014"])
for row in data:
    ws.append(row)

ft=Font(bold=True)
for row in ws["A1:K1"]:
    for cell in row:
        cell.font = ft


# tab = Table(displayName="Table1", ref="A1:E5")
# ws.add_table(tab)
wb.save('sample.xlsx')